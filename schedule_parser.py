import json
import logging
import os
import sys
import time
from datetime import datetime

import openpyxl

from models import Shift

logger = logging.getLogger(__name__)

# Row in the XLSX that contains column headers (rows 1-5 are title block)
HEADER_ROW = 6

# Required column headers — must match XLSX exactly (case-sensitive)
HDR_DATE     = "Дата"
HDR_DAY_TYPE = "Day-type"

DEPARTMENT_HEADERS: list[str] = [
    "Приймальне відділення",
    "Анестезіологія",
    "реанімація",
    "хірургія",
    "акушерство",
    "травматологія",
    "неврологія",
    "УЗД",
    "Дитяче відділення",
]

# Present in XLSX but out of POC scope — parser logs INFO and skips it
HDR_URGENCIA = "Ургенція спеціалістів на дому"

REQUIRED_HEADERS: list[str] = [HDR_DATE, HDR_DAY_TYPE] + DEPARTMENT_HEADERS

VALID_DAY_TYPES = {"labor", "holiday", "other"}


def check_file_freshness(xlsx_path: str, threshold_seconds: int = 60) -> None:
    mtime = os.path.getmtime(xlsx_path)
    age = time.time() - mtime
    if age < threshold_seconds:
        modified_at = datetime.fromtimestamp(mtime).strftime("%Y-%m-%d %H:%M:%S")
        raise RuntimeError(
            f"XLSX was modified {age:.0f}s ago (at {modified_at}) — "
            f"possible upload in progress. Path: {xlsx_path}"
        )


def _load_contacts(contacts_path: str) -> dict[str, dict]:
    try:
        with open(contacts_path, encoding="utf-8") as f:
            entries = json.load(f)
    except OSError as exc:
        logger.critical("Cannot open contacts file: %s — %s", contacts_path, exc)
        print(f"[CONTACTS] ❌ cannot open: {contacts_path}")
        sys.exit(1)
    except json.JSONDecodeError as exc:
        logger.critical("Invalid JSON in contacts file: %s — %s", contacts_path, exc)
        print(f"[CONTACTS] ❌ invalid JSON: {contacts_path}")
        sys.exit(1)

    contacts: dict[str, dict] = {}
    for entry in entries:
        name = str(entry.get("name", "")).strip()
        if name:
            contacts[name] = entry
    return contacts


def _build_header_map(sheet, header_row: int) -> dict[str, int]:
    rows = list(sheet.iter_rows(min_row=header_row, max_row=header_row, values_only=True))
    header_map: dict[str, int] = {}
    for col_idx, value in enumerate(rows[0], start=1):
        if value is not None:
            header_map[str(value).strip()] = col_idx
    return header_map


def _require_headers(header_map: dict[str, int], required: list[str]) -> None:
    missing = [h for h in required if h not in header_map]
    if not missing:
        return
    for h in missing:
        logger.critical("Missing required XLSX header: '%s'", h)
        print(f"[XLSX]    ❌ missing required header: '{h}'")
    sys.exit(1)


def _parse_date(raw: object, context: str) -> str | None:
    if isinstance(raw, datetime):
        return raw.strftime("%Y-%m-%d")
    try:
        return datetime.strptime(str(raw).strip(), "%d-%m-%Y").strftime("%Y-%m-%d")
    except ValueError:
        logger.warning("Unrecognised date '%s' on row %s — skipping", raw, context)
        return None


def parse_schedule(xlsx_path: str, contacts_path: str) -> list[Shift]:
    check_file_freshness(xlsx_path)

    try:
        wb = openpyxl.load_workbook(xlsx_path, data_only=True)
    except Exception as exc:
        logger.critical("Cannot open XLSX: %s — %s", xlsx_path, exc)
        print(f"[XLSX]    ❌ cannot open: {xlsx_path}")
        sys.exit(1)

    ws = wb.worksheets[0]
    header_map = _build_header_map(ws, HEADER_ROW)
    _require_headers(header_map, REQUIRED_HEADERS)

    if HDR_URGENCIA in header_map:
        logger.info("Column '%s' found but skipped (out of POC scope)", HDR_URGENCIA)

    contacts = _load_contacts(contacts_path)
    shifts: list[Shift] = []
    skip_count = 0

    for row in ws.iter_rows(min_row=HEADER_ROW + 1, values_only=True):
        raw_date     = row[header_map[HDR_DATE] - 1]
        raw_day_type = row[header_map[HDR_DAY_TYPE] - 1]

        if raw_date is None:
            continue

        shift_date = _parse_date(raw_date, str(raw_date))
        if shift_date is None:
            continue

        day_type = str(raw_day_type or "").strip().lower()
        if day_type not in VALID_DAY_TYPES:
            logger.warning("Unknown day_type '%s' on %s — skipping row", day_type, shift_date)
            continue

        for dept in DEPARTMENT_HEADERS:
            if dept not in header_map:
                continue
            cell_value = row[header_map[dept] - 1]
            if cell_value is None:
                continue

            employee_name = str(cell_value).strip()
            if not employee_name:
                continue

            if employee_name not in contacts:
                logger.warning("'%s' not in contacts.json — skipping", employee_name)
                skip_count += 1
                continue

            contact = contacts[employee_name]
            primary = str(contact.get("primary_channel", "telegram")).strip()
            channel_id = str(contact.get("channels", {}).get(primary, "")).strip()

            if not channel_id:
                logger.warning("Empty contact_id for '%s' — skipping", employee_name)
                skip_count += 1
                continue

            shifts.append(Shift(
                employee_name=employee_name,
                department=dept,
                day_type=day_type,
                shift_date=shift_date,
                messenger=primary,
                contact_id=channel_id,
            ))

    if skip_count:
        logger.warning("Total skipped (no contact match or missing contact_id): %d", skip_count)

    return shifts
